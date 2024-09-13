import csv
import datetime
import logging
import re
import zipfile
from collections import namedtuple
from decimal import Decimal
from functools import reduce
from io import BytesIO, StringIO
from numbers import Number

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.db.models import Avg, Count, Sum, Max, Min
from django.db.models.fields.related_descriptors import ManyToManyDescriptor
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

from .utils import (
    get_relation_fields_from_model,
    get_properties_from_model,
    get_direct_fields_from_model,
    get_model_from_path_string,
    get_custom_fields_from_model,
)

DisplayField = namedtuple(
    "DisplayField",
    "path path_verbose field field_verbose aggregate total group choices field_type",
)

User = get_user_model()


def generate_filename(title, ends_with):
    title = title.split('.')[0]
    title.replace(' ', '_')
    title += ('_' + datetime.datetime.now().strftime("%m%d_%H%M"))
    if not title.endswith(ends_with):
        title += ends_with
    return title


class DataExportMixin(object):
    max_rows = 10000

    def build_sheet(self, data, ws, sheet_name='report', header=None, widths=None):
        first_row = 1
        column_base = 1

        ws.title = re.sub(r'\W+', '', sheet_name)[:30]
        if header:
            for i, header_cell in enumerate(header):
                cell = ws.cell(row=first_row, column=i + column_base)
                cell.value = header_cell
                cell.font = Font(bold=True)
                if widths:
                    ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

        for row in data:
            for i in range(len(row)):
                item = row[i]
                if isinstance(item, str):
                    try:
                        row[i] = str(item)
                    except UnicodeDecodeError:
                        row[i] = str(item.decode('utf-8', 'ignore'))
                elif isinstance(item, dict):
                    row[i] = str(item)
                elif not isinstance(item, (int, float, bool)):
                    row[i] = str(item)
            try:
                ws.append(row)
            except ValueError as e:
                ws.append([str(e)])
            except Exception:
                ws.append(['Unknown Error'])

    def build_xlsx_response(self, wb, title="report"):
        title = generate_filename(title, '.xlsx')
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={title}'
        return response

    def build_csv_response(self, csv_content, title="report"):
        title = generate_filename(title, '.csv')
        response = HttpResponse(
            csv_content,
            content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={title}'
        return response

    def build_zip_response(self, files, title="report"):
        logger = logging.getLogger(__name__)

        logger.debug(f"Building zip response with files: {list(files.keys())}")

        title = generate_filename(title, '.zip')
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_name, file_content in files.items():
                if file_name.endswith('.csv'):
                    zipf.writestr(file_name, file_content.decode('utf-8'))
                else:
                    zipf.writestr(file_name, file_content)

        zip_buffer.seek(0)

        response = HttpResponse(
            zip_buffer,
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename="{title}"'

        zip_buffer.close()

        logger.debug(f"Zip response prepared for download: {title}")

        return response

    def list_to_workbook(self, data, title='report', header=None, widths=None):
        logger.info("Creating Workbook...")
        wb = Workbook()
        title = re.sub(r'\W+', '', title)[:30]

        if isinstance(data, dict):
            i = 0
            for sheet_name, sheet_data in data.items():
                logger.info(f"Creating sheet: {sheet_name}")
                if i > 0:
                    wb.create_sheet()
                ws = wb.worksheets[i]
                self.build_sheet(
                    sheet_data, ws, sheet_name=sheet_name, header=header)
                i += 1
        else:
            logger.info("Creating default sheet")
            ws = wb.worksheets[0]
            self.build_sheet(data, ws, header=header, widths=widths)
        logger.info("Workbook created successfully")
        return wb

    def list_to_xlsx_file(self, data, title='report', header=None, widths=None):
        logger.info("Generating XLSX file...")
        wb = self.list_to_workbook(data, title, header, widths)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info("XLSX file generated successfully")
        return output.getvalue()

    def list_to_csv_file(self, data, title='report', header=None, widths=None):
        logger.info("Generating CSV file...")
        output = StringIO()
        sh = self.list_to_workbook(data, title, header, widths).active
        c = csv.writer(output)
        if header:
            logger.info("Writing header to CSV")
            c.writerow(header)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
        output.seek(0)
        logger.info("CSV file generated successfully")
        return output.getvalue()

    def list_to_xlsx_response(self, data, title='report', header=None, widths=None):
        total_rows = len(data)
        logger.info(f"Total Rows: {total_rows}, Max Rows: {self.max_rows}")
        if total_rows > self.max_rows:
            logger.info("Total rows exceed max_rows, generating ZIP response")
            return self.list_to_zip_response(data, title, header, widths, file_type="xlsx")
        else:
            logger.info("Generating XLSX response")
            wb = self.list_to_workbook(data, title, header, widths)
            return self.build_xlsx_response(wb, title=title)

    def list_to_csv_response(self, data, title='report', header=None, widths=None):
        total_rows = len(data)
        logger.info(f"Total Rows: {total_rows}, Max Rows: {self.max_rows}")
        if total_rows > self.max_rows:
            logger.info("Total rows exceed max_rows, generating ZIP response")
            return self.list_to_zip_response(data, title, header, widths, file_type="csv")
        else:
            logger.info("Generating CSV response")
            csv_content = self.list_to_csv_file(data, title, header, widths)
            return self.build_csv_response(csv_content, title)

    def list_to_zip_response(self, data, title="report", header=None, widths=None, file_type="xlsx"):
        logger = logging.getLogger(__name__)

        total_rows = len(data)
        num_parts = (total_rows // self.max_rows) + (1 if total_rows % self.max_rows != 0 else 0)

        logger.info(f"Total rows: {total_rows}, Max rows per file: {self.max_rows}, Number of parts: {num_parts}")

        files = {}
        for part in range(num_parts):
            start_row = part * self.max_rows
            end_row = min(start_row + self.max_rows, total_rows)
            part_data = data[start_row:end_row]
            part_title = f"{title}_part{part + 1}"

            logger.info(f"Processing part {part + 1}: Rows {start_row} to {end_row}")

            if file_type == "csv":
                csv_content = self.list_to_csv_file(part_data, part_title, header, widths)
                files[f"{part_title}.csv"] = csv_content.encode('utf-8')
            elif file_type == "xlsx":
                xlsx_content = self.list_to_xlsx_file(part_data, part_title, header, widths)
                files[f"{part_title}.xlsx"] = xlsx_content

        logger.info(f"Files to add to zip: {files.keys()}")

        return self.build_zip_response(files, title)

    def add_aggregates(self, queryset, display_fields):
        agg_funcs = {
            'Avg': Avg, 'Min': Min, 'Max': Max, 'Count': Count, 'Sum': Sum
        }

        for display_field in display_fields:
            if display_field.aggregate:
                func = agg_funcs[display_field.aggregate]
                full_name = display_field.path + display_field.field
                queryset = queryset.annotate(func(full_name))

        return queryset

    def report_to_list(self, queryset, display_fields, user=None, property_filters=[], preview=False):
        model_class = queryset.model

        def can_change_or_view(model):
            if user is None:
                return True
            model_name = model._meta.model_name
            app_label = model._meta.app_label
            can_change = user.has_perm(app_label + '.change_' + model_name)
            can_view = user.has_perm(app_label + '.view_' + model_name)

            return can_change or can_view

        if not can_change_or_view(model_class):
            return [], 'Permission Denied'

        queryset = queryset.inplace()

        if isinstance(display_fields, list):
            new_display_fields = []
            for display_field in display_fields:
                field_list = display_field.split('__')
                field = field_list[-1]
                path = '__'.join(field_list[:-1])

                if path:
                    path += '__'

                new_model = get_model_from_path_string(model_class, path)
                model_field = new_model._meta.get_field(field)
                choices = model_field.choices
                new_display_fields.append(DisplayField(
                    path, '', field, '', '', None, None, choices, ''
                ))

            display_fields = new_display_fields

        group = [df.path + df.field for df in display_fields if df.group]
        if group:
            for field in display_fields:
                if (not field.group) and (not field.aggregate):
                    field.aggregate = 'Max'

        message = ""
        objects = self.add_aggregates(queryset, display_fields)

        display_field_paths = []
        property_list = {}
        custom_list = {}
        display_totals = {}

        for i, display_field in enumerate(display_fields):
            model = get_model_from_path_string(model_class, display_field.path)

            if display_field.field_type == "Invalid":
                continue

            if not model or can_change_or_view(model):
                display_field_key = display_field.path + display_field.field

                if display_field.field_type == "Property":
                    property_list[i] = display_field_key
                elif display_field.field_type == "Custom Field":
                    custom_list[i] = display_field_key
                elif display_field.aggregate == "Avg":
                    display_field_key += '__avg'
                elif display_field.aggregate == "Max":
                    display_field_key += '__max'
                elif display_field.aggregate == "Min":
                    display_field_key += '__min'
                elif display_field.aggregate == "Count":
                    display_field_key += '__count'
                elif display_field.aggregate == "Sum":
                    display_field_key += '__sum'

                if display_field.field_type not in ('Property', 'Custom Field'):
                    display_field_paths.append(display_field_key)

                if display_field.total:
                    display_totals[display_field_key] = Decimal(0)

            else:
                message += 'Error: Permission denied on access to {0}.'.format(
                    display_field.name
                )

        def increment_total(display_field_key, val):
            if display_field_key in display_totals:
                if isinstance(val, bool):
                    display_totals[display_field_key] += Decimal(val)
                elif isinstance(val, Number):
                    display_totals[display_field_key] += Decimal(str(val))
                elif val:
                    display_totals[display_field_key] += Decimal(1)

        if not group:
            display_field_paths.insert(0, 'pk')

            m2m_relations = []
            for position, property_path in property_list.items():
                property_root = property_path.split('__')[0]
                root_class = model_class

                try:
                    property_root_class = getattr(root_class, property_root)
                except AttributeError:
                    continue

                if type(property_root_class) == ManyToManyDescriptor:
                    display_field_paths.insert(1, '%s__pk' % property_root)
                    m2m_relations.append(property_root)

        if group:
            values = objects.values(*group)
            values = self.add_aggregates(values, display_fields)
            filtered_report_rows = [
                [row[field] for field in display_field_paths]
                for row in values
            ]
            for row in filtered_report_rows:
                for pos, field in enumerate(display_field_paths):
                    increment_total(field, row[pos])
        else:
            filtered_report_rows = []
            values_and_properties_list = []

            values_list = objects.values_list(*display_field_paths)

            for row in values_list:
                row = list(row)
                values_and_properties_list.append(row[1:])
                obj = None
                remove_row = False
                for property_filter in property_filters:
                    if not obj:
                        obj = model_class.objects.get(pk=row.pop(0))
                    root_relation = property_filter.path.split('__')[0]
                    if root_relation in m2m_relations:
                        pk = row[0]
                        if pk is not None:
                            m2m_obj = getattr(obj, root_relation).get(pk=pk)
                            val = reduce(getattr, [property_filter.field], m2m_obj)
                        else:
                            val = None
                    else:
                        if property_filter.field_type == 'Custom Field':
                            for relation in property_filter.path.split('__'):
                                if hasattr(obj, root_relation):
                                    obj = getattr(obj, root_relation)
                            val = obj.get_custom_value(property_filter.field)
                        else:
                            val = reduce(getattr, (property_filter.path + property_filter.field).split('__'), obj)
                    if property_filter.filter_property(val):
                        remove_row = True
                        values_and_properties_list.pop()
                        break
                if not remove_row:
                    for i, field in enumerate(display_field_paths[1:]):
                        increment_total(field, row[i + 1])

                    for position, display_property in property_list.items():
                        if not obj:
                            obj = model_class.objects.get(pk=row.pop(0))
                        relations = display_property.split('__')
                        root_relation = relations[0]
                        if root_relation in m2m_relations:
                            pk = row.pop(0)
                            if pk is not None:
                                m2m_obj = getattr(obj, root_relation).get(pk=pk)
                                val = reduce(getattr, relations[1:], m2m_obj)
                            else:
                                val = None
                        else:
                            try:
                                val = reduce(getattr, relations, obj)
                            except AttributeError:
                                val = None
                            values_and_properties_list[-1].insert(position, val)
                            increment_total(display_property, val)

                    for position, display_custom in custom_list.items():
                        if not obj:
                            obj = model_class.objects.get(pk=row.pop(0))
                        val = obj.get_custom_value(display_custom)
                        values_and_properties_list[-1].insert(position, val)
                        increment_total(display_custom, val)

                    filtered_report_rows.append(values_and_properties_list[-1])

                if preview and len(filtered_report_rows) == 50:
                    break

        if hasattr(display_fields, 'filter'):
            defaults = {
                None: str,
                datetime.date: lambda: datetime.date(datetime.MINYEAR, 1, 1),
                datetime.datetime: lambda: datetime.datetime(datetime.MINYEAR, 1, 1),
            }

            sort_fields = display_fields.filter(sort__gt=0).order_by('-sort')
            sort_values = sort_fields.values_list('position', 'sort_reverse')

            for pos, reverse in sort_values:
                column = (row[pos] for row in filtered_report_rows)
                type_col = (type(val) for val in column if val is not None)
                field_type = next(type_col, None)
                default = defaults.get(field_type, field_type)()

                filtered_report_rows = sorted(
                    filtered_report_rows,
                    key=lambda row: self.sort_helper(row[pos], default),
                    reverse=reverse,
                )

        values_and_properties_list = filtered_report_rows

        choice_lists = {}
        for df in display_fields:
            if df.choices and hasattr(df, 'choices_dict'):
                df_choices = df.choices_dict
                df_choices[''] = ''
                df_choices[None] = ''
                choice_lists[df.position] = df_choices

        display_formats = {}

        for df in display_fields:
            if hasattr(df, 'display_format') and df.display_format:
                display_formats[df.position] = df.display_format

        def formatter(value, style):
            try:
                value = Decimal(value)
            except Exception:
                pass

            try:
                return style.string.format(value)
            except ValueError:
                return value

        final_list = []

        for row in values_and_properties_list:
            row = list(row)

            for position, choice_list in choice_lists.items():
                try:
                    row[position] = str(choice_list[row[position]])
                except Exception:
                    row[position] = str(row[position])

            for pos, style in display_formats.items():
                row[pos] = formatter(row[pos], style)

            final_list.append(row)

        values_and_properties_list = final_list

        if display_totals:
            display_totals_row = []

            fields_and_properties = list(display_field_paths[0 if group else 1:])

            for position, value in property_list.items():
                fields_and_properties.insert(position, value)

            for field in fields_and_properties:
                display_totals_row.append(display_totals.get(field, ''))

            for pos, style in display_formats.items():
                display_totals_row[pos] = formatter(display_totals_row[pos], style)

            values_and_properties_list.append(
                ['TOTALS'] + (len(fields_and_properties) - 1) * ['']
            )
            values_and_properties_list.append(display_totals_row)

        return values_and_properties_list, message


class GetFieldsMixin(object):
    def get_fields(self, model_class, field_name='', path='', path_verbose=''):
        fields = get_direct_fields_from_model(model_class)
        properties = get_properties_from_model(model_class)
        custom_fields = get_custom_fields_from_model(model_class)
        app_label = model_class._meta.app_label
        model = model_class

        if field_name != '':
            field = model_class._meta.get_field(field_name)
            direct = field.concrete
            if path_verbose:
                path_verbose += "::"
            if field.many_to_many and hasattr(field, 'm2m_reverse_field_name'):
                path_verbose += field.m2m_reverse_field_name()
            else:
                path_verbose += field.name

            path += field_name
            path += '__'
            if direct:
                new_model = field.related_model
                path_verbose = new_model.__name__.lower()
            else:
                new_model = field.related_model
                path_verbose = new_model.__name__.lower()

            fields = get_direct_fields_from_model(new_model)

            custom_fields = get_custom_fields_from_model(new_model)
            properties = get_properties_from_model(new_model)
            app_label = new_model._meta.app_label
            model = new_model

        return {
            'fields': fields,
            'custom_fields': custom_fields,
            'properties': properties,
            'path': path,
            'path_verbose': path_verbose,
            'app_label': app_label,
            'model': model,
        }

    def get_related_fields(self, model_class, field_name, path="", path_verbose=""):
        if field_name:
            field = model_class._meta.get_field(field_name)
            direct = field.concrete
            if direct:
                try:
                    related_field = field.remote_field
                except AttributeError:
                    related_field = field.related
                try:
                    new_model = related_field.parent_model()
                except AttributeError:
                    new_model = related_field.model
            else:
                new_model = field.related_model

            if path_verbose:
                path_verbose += "::"
            path_verbose += field.name

            path += field_name
            path += '__'
        else:
            new_model = model_class

        new_fields = get_relation_fields_from_model(new_model)
        model_ct = ContentType.objects.get_for_model(new_model)

        return (new_fields, model_ct, path)
